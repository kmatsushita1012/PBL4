import openpyxl as px

def input():
	
	wb = px.load_workbook("input.xlsx")

	ws = wb.active

	info = {}
	
	table = []
	for i in range(ws.cell(3, 2).value):
		row = []
		for j in range(ws.cell(4, 2).value):
			row.append(ws.cell(2 + i, 5 + j).value)
		table.append(row)
	info["table"] = table

	info["class"] = ws.cell(1, 2).value
	info["num"] = ws.cell(2, 2).value

	info["teacher"] = ws.cell(5, 2).value
	info["mail"] = ws.cell(6, 2).value

	info["start time"] = ws.cell(7, 2).value
	info["end time"] = ws.cell(8, 2).value

	wb.close()

	return info

def main():
	
	info = input()
	
	print(info)

if __name__ == "__main__":
	main()